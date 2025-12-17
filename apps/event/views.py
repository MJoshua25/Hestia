from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy, reverse
from django.shortcuts import get_object_or_404, redirect
from .models import Event, Commission
from .forms import EventForm, CommissionForm
from apps.member.models import Member
from core.mixins import AdminOrDelegateRequiredMixin, AdminRequiredMixin
from django.views import View
from django.http import JsonResponse, HttpResponse
from django.views.generic import TemplateView
from django.db.models import Prefetch, Count
import json
from .services import AssignmentService
from .models import Assignment
from apps.member.models import Member
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

class EventListView(LoginRequiredMixin, ListView):
    model = Event
    template_name = 'event/event_list.html'
    context_object_name = 'events'
    paginate_by = 20

    def get_queryset(self):
        return Event.objects.all().order_by('date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Separate past and future events?
        # PRD says: "voir deux sections : Événements à venir et Événements passés"
        # But ListView usually iterates over one list.
        # I can split them in template or context.
        from django.utils import timezone
        now = timezone.now()
        context['upcoming_events'] = Event.objects.filter(date__gte=now).order_by('date')
        context['past_events'] = Event.objects.filter(date__lt=now).order_by('-date')
        context['undated_events'] = Event.objects.filter(date__isnull=True)
        return context

class EventDetailView(LoginRequiredMixin, DetailView):
    model = Event
    template_name = 'event/event_detail.html'
    context_object_name = 'event'

class EventCreateView(AdminOrDelegateRequiredMixin, CreateView):
    model = Event
    form_class = EventForm
    template_name = 'event/event_form.html'
    
    def get_success_url(self):
        return reverse('event_detail', kwargs={'pk': self.object.pk})

class EventUpdateView(AdminOrDelegateRequiredMixin, UpdateView):
    model = Event
    form_class = EventForm
    template_name = 'event/event_form.html'
    
    def get_success_url(self):
        return reverse('event_detail', kwargs={'pk': self.object.pk})

class EventDeleteView(AdminRequiredMixin, DeleteView):
    model = Event
    template_name = 'event/event_confirm_delete.html'
    success_url = reverse_lazy('event_list')

# Commission Management
class CommissionCreateView(AdminOrDelegateRequiredMixin, CreateView):
    model = Commission
    form_class = CommissionForm
    template_name = 'event/commission_form.html'

    def form_valid(self, form):
        event = get_object_or_404(Event, pk=self.kwargs['event_id'])
        form.instance.event = event
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('event_detail', kwargs={'pk': self.kwargs['event_id']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = get_object_or_404(Event, pk=self.kwargs['event_id'])
        return context

class CommissionUpdateView(AdminOrDelegateRequiredMixin, UpdateView):
    model = Commission
    form_class = CommissionForm
    template_name = 'event/commission_form.html'

    def get_success_url(self):
        return reverse('event_detail', kwargs={'pk': self.object.event.pk})

class CommissionDeleteView(AdminOrDelegateRequiredMixin, DeleteView):
    model = Commission
    template_name = 'event/commission_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('event_detail', kwargs={'pk': self.object.event.pk})

# Commission Attribution & Management UI
class CommissionManageView(AdminOrDelegateRequiredMixin, TemplateView):
    template_name = 'event/commission_manage.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event = get_object_or_404(Event, pk=self.kwargs['event_id'])
        context['event'] = event
        return context

# JSON API for Vue.js
class CommissionDataAPI(AdminOrDelegateRequiredMixin, View):
    def get(self, request, event_id):
        event = get_object_or_404(Event, pk=event_id)
        
        # 1. Get commissions with assignments
        commissions = event.commissions.prefetch_related(
            'assignments__member'
        ).all()
        
        commissions_data = []
        assigned_member_ids = set()
        
        for c in commissions:
            assignments = []
            for a in c.assignments.all():
                assignments.append({
                    'id': a.id,
                    'member_id': a.member.id,
                    'member_name': str(a.member),
                    'member_room': a.member.room_number,
                    'member_phone': a.member.phone_number
                })
                assigned_member_ids.add(a.member.id)
                
            commissions_data.append({
                'id': c.id,
                'name': c.name,
                'min': c.min_capacity,
                'max': c.max_capacity,
                'responsible': str(c.responsible) if c.responsible else None,
                'assignments': assignments,
                'current_count': len(assignments),
                'is_full': c.is_full
            })
            
        # 2. Get available members (not assigned yet)
        all_members = Member.objects.filter(statut=True).order_by('last_name')
        members_data = []
        for m in all_members:
            members_data.append({
                'id': m.id,
                'name': str(m),
                'room': m.room_number,
                'phone': m.phone_number,
                'is_assigned': m.id in assigned_member_ids,
                'assigned_to_commission_id': self._get_commission_id(commissions_data, m.id) if m.id in assigned_member_ids else None
            })
            
        return JsonResponse({
            'commissions': commissions_data,
            'members': members_data
        })
        
    def _get_commission_id(self, commissions_data, member_id):
        for c in commissions_data:
            for a in c['assignments']:
                if a['member_id'] == member_id:
                    return c['id']
        return None

class AssignmentAutoAPI(AdminOrDelegateRequiredMixin, View):
    def post(self, request, event_id):
        try:
            data = json.loads(request.body)
            member_ids = data.get('member_ids', [])
            event = get_object_or_404(Event, pk=event_id)
            
            result = AssignmentService.assign_automatically(event, member_ids, request.user)
            return JsonResponse(result)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

class AssignmentManualAPI(AdminOrDelegateRequiredMixin, View):
    def post(self, request, event_id):
        try:
            data = json.loads(request.body)
            commission_id = data.get('commission_id')
            member_id = data.get('member_id')
            action = data.get('action') # 'add', 'move', 'remove'
            
            event = get_object_or_404(Event, pk=event_id)
            commission = get_object_or_404(Commission, pk=commission_id, event=event) if commission_id else None
            member = get_object_or_404(Member, pk=member_id)
            
            if action == 'add':
                # Check constraints
                if Assignment.objects.filter(commission__event=event, member=member).exists():
                    return JsonResponse({'status': 'error', 'message': "Ce membre est déjà assigné une commission"}, status=400)
                
                if commission.is_full:
                    return JsonResponse({'status': 'error', 'message': "La commission est pleine"}, status=400)
                    
                Assignment.objects.create(commission=commission, member=member, assigned_by=request.user)
                
            elif action == 'remove':
                # Find assignment for this member in this event (or specific commission)
                 Assignment.objects.filter(commission__event=event, member=member).delete()
                 
            elif action == 'move':
                if not commission:
                     return JsonResponse({'status': 'error', 'message': "Commission cible invalide"}, status=400)
                
                if commission.is_full:
                    return JsonResponse({'status': 'error', 'message': "La commission cible est pleine"}, status=400)
                    
                # Atomic update
                with transaction.atomic():
                    Assignment.objects.filter(commission__event=event, member=member).delete()
                    Assignment.objects.create(commission=commission, member=member, assigned_by=request.user)
            
            return JsonResponse({'status': 'success'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

# Export Views
def export_commissions_excel(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    commissions = event.commissions.prefetch_related('assignments__member')
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
        
    for commission in commissions:
        ws = wb.create_sheet(title=commission.name[:30]) # Sheet name limit 31 chars
        
        # Header
        headers = ['Prénom', 'Nom', 'Chambre', 'Téléphone']
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            
        assignments = commission.assignments.all().order_by('member__last_name')
        for assignment in assignments:
            m = assignment.member
            ws.append([m.first_name, m.last_name, m.room_number, m.phone_number])
            
        # Add Responsable info if exists
        if commission.responsible:
            ws.append([])
            ws.append(['Responsable :', str(commission.responsible)])
            
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{event.title}_commissions.xlsx"'
    wb.save(response)
    return response

def export_commissions_pdf(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    commissions = event.commissions.prefetch_related('assignments__member').order_by('name')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{event.title}_commissions.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    h2_style = styles['Heading2']
    normal_style = styles['Normal']

    # Title
    elements.append(Paragraph(f"{event.title}", title_style))
    elements.append(Paragraph(f"Commissions & Attributions", h2_style))
    elements.append(Spacer(1, 12))

    for commission in commissions:
        elements.append(Paragraph(f"Commission: {commission.name}", h2_style))
        if commission.responsible:
             elements.append(Paragraph(f"Responsable: {commission.responsible}", normal_style))
        
        elements.append(Spacer(1, 6))
        
        data = [['Nom', 'Chambre', 'Téléphone']]
        assignments = commission.assignments.all().order_by('member__last_name')
        
        if not assignments:
             elements.append(Paragraph("Aucun membre assigné.", normal_style))
        else:
            for assignment in assignments:
                m = assignment.member
                data.append([f"{m.first_name} {m.last_name}", m.room_number, m.phone_number])
                
            table = Table(data, colWidths=[250, 80, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
        elements.append(Spacer(1, 20))

    doc.build(elements)
    return response
